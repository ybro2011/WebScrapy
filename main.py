from gevent import monkey
monkey.patch_all()

import os
import math
import re
import gc
import logging
import time
from datetime import datetime
from flask import Flask, render_template, request, Response, stream_with_context, send_file, jsonify
import googlemaps
from openpyxl import Workbook
from dotenv import load_dotenv
import json
from openpyxl.styles import Font, PatternFill

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Constants for grid search
EARTH_RADIUS = 6371  # Earth's radius in kilometers
DEFAULT_GRID_RADIUS = 5  # Default 5km radius
GRID_SIZE_LOW = 3
GRID_SIZE_MEDIUM = 5
GRID_SIZE_HIGH = 7

# Ensure the Documents directory exists
DOCUMENTS_DIR = os.path.join(os.path.expanduser('~'), 'Documents')
os.makedirs(DOCUMENTS_DIR, exist_ok=True)

# Global variables for task tracking
active_tasks = {}
task_results = {}

def get_grid_size(density):
    """
    Get the grid size based on density setting.
    Returns the number of points in one dimension of the grid.
    """
    if density == 'low':
        return GRID_SIZE_LOW
    elif density == 'high':
        return GRID_SIZE_HIGH
    else:  # medium
        return GRID_SIZE_MEDIUM

def create_search_grid(center_lat, center_lng, radius, density):
    """
    Create a grid of search points around a center point.
    Returns a list of (lat, lng) tuples including the center point.
    """
    grid_points = []
    
    # Get grid size based on density
    grid_size = get_grid_size(density)
    half_size = (grid_size - 1) // 2
    
    # Calculate the distance between grid points (in radians)
    distance_radians = radius / EARTH_RADIUS
    
    # Convert center coordinates to radians
    center_lat_rad = math.radians(center_lat)
    center_lng_rad = math.radians(center_lng)
    
    # Calculate the distance between grid points in degrees
    # Rough approximation: 1 degree ≈ 111 km
    distance_degrees = (radius * 2) / (grid_size * 111)
    
    # Calculate the starting point (top-left of the grid)
    start_lat = center_lat + (radius / 111)
    start_lng = center_lng - (radius / 111)
    
    # Create grid points
    for i in range(-half_size, half_size + 1):
        for j in range(-half_size, half_size + 1):
            lat = start_lat - (i * distance_degrees)
            lng = start_lng + (j * distance_degrees)
            
            # Calculate new latitude
            new_lat_rad = math.asin(
                math.sin(center_lat_rad) * math.cos(distance_radians) +
                math.cos(center_lat_rad) * math.sin(distance_radians) * math.cos(math.radians(i * 90))
            )
            
            # Calculate new longitude
            new_lng_rad = center_lng_rad + math.atan2(
                math.sin(math.radians(j * 90)) * math.sin(distance_radians) * math.cos(center_lat_rad),
                math.cos(distance_radians) - math.sin(center_lat_rad) * math.sin(new_lat_rad)
            )
            
            # Convert back to degrees
            new_lat = math.degrees(new_lat_rad)
            new_lng = math.degrees(new_lng_rad)
            
            grid_points.append((new_lat, new_lng))
    
    return grid_points

def get_location_coordinates(gmaps, location_name):
    """
    Get coordinates for a location name using Google Maps Geocoding API.
    Returns a tuple of (latitude, longitude).
    """
    try:
        geocode_result = gmaps.geocode(location_name)
        if not geocode_result:
            raise ValueError(f"Could not find coordinates for location: {location_name}")
        
        location = geocode_result[0]['geometry']['location']
        return (location['lat'], location['lng'])
    except Exception as e:
        logger.error(f"Error getting coordinates for {location_name}: {e}")
        raise

def search_places(gmaps, location, query, radius=5000):
    """Search for places near a location using Google Maps Places API."""
    try:
        logger.info(f"Starting search at location {location} with radius {radius}")
        all_results = []
        
        # Initial search
        places_result = gmaps.places_nearby(
            location=location,
            radius=radius,
            keyword=query,
            type='establishment'
        )
        
        if 'results' in places_result:
            all_results.extend(places_result['results'])
            logger.info(f"Initial search found {len(places_result['results'])} results")
            
            # Handle pagination with more conservative rate limiting
            while 'next_page_token' in places_result and len(all_results) < 60:
                logger.info(f"Waiting for next page token (page {len(all_results)//20 + 1})")
                time.sleep(5)  # Increased delay to 5 seconds
                
                try:
                    places_result = gmaps.places_nearby(
                        location=location,
                        radius=radius,
                        keyword=query,
                        type='establishment',
                        page_token=places_result['next_page_token']
                    )
                    
                    if 'results' in places_result:
                        all_results.extend(places_result['results'])
                        logger.info(f"Page {len(all_results)//20 + 1} found {len(places_result['results'])} results")
                        logger.info(f"Total results found: {len(all_results)}")
                except Exception as e:
                    logger.error(f"Error during pagination: {str(e)}")
                    break
        
        return all_results
    except Exception as e:
        logger.error(f"Error in search_places: {str(e)}")
        return []

def get_place_details(gmaps, place_id):
    """
    Get detailed information about a place using Google Maps Places API.
    Returns a dictionary of place details.
    """
    try:
        place_details = gmaps.place(place_id, fields=['name', 'formatted_address', 'formatted_phone_number', 'website', 'rating', 'user_ratings_total'])
        result = place_details.get('result', {})
        
        return {
            'name': result.get('name', ''),
            'address': result.get('formatted_address', ''),
            'phone': result.get('formatted_phone_number', ''),
            'website': result.get('website', ''),
            'rating': result.get('rating', ''),
            'reviews': result.get('user_ratings_total', ''),
            'email': ''  # Empty email since we're not scraping websites
        }
    except Exception as e:
        logger.error(f"Error getting place details: {e}")
        return {
            'name': '',
            'address': '',
            'phone': '',
            'website': '',
            'rating': '',
            'reviews': '',
            'email': ''
        }

def save_to_excel(businesses, filename):
    """
    Save business data to an Excel file.
    Returns the path to the saved file.
    """
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = ['Name', 'Address', 'Phone', 'Website', 'Rating', 'Reviews']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, business in enumerate(businesses, 2):
        ws.cell(row=row, column=1, value=business['name'])
        ws.cell(row=row, column=2, value=business['address'])
        ws.cell(row=row, column=3, value=business['phone'])
        ws.cell(row=row, column=4, value=business['website'])
        ws.cell(row=row, column=5, value=business['rating'])
        ws.cell(row=row, column=6, value=business['reviews'])

    
    # Save file
    file_path = os.path.join(DOCUMENTS_DIR, filename)
    wb.save(file_path)
    return file_path

def save_checkpoint(data, filename):
    """Save checkpoint data to a file"""
    try:
        with open(filename, 'w') as f:
            json.dump(data, f)
    except Exception as e:
        logger.error(f"Error saving checkpoint: {e}")

def load_checkpoint(filename):
    """Load checkpoint data from a file"""
    try:
        if os.path.exists(filename):
            with open(filename, 'r') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Error loading checkpoint: {e}")
    return None

def cleanup_checkpoint(filename):
    """Remove checkpoint file after successful completion"""
    try:
        if os.path.exists(filename):
            os.remove(filename)
    except Exception as e:
        logger.error(f"Error cleaning up checkpoint: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])
def search():
    try:
        api_key = request.form['api_key']
        location_name = request.form['location']
        industry = request.form['industry']
        radius = float(request.form.get('radius', DEFAULT_GRID_RADIUS))
        density = request.form.get('density', 'medium')
        
        # Generate a unique task ID
        task_id = f"{industry}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Initialize the task
        active_tasks[task_id] = {
            'status': 'running',
            'progress': 0,
            'message': 'Starting search...'
        }
        
        def generate_updates():
            try:
                # Try to load checkpoint
                checkpoint_file = f"checkpoint_{task_id}.json"
                checkpoint = load_checkpoint(checkpoint_file)
                
                if checkpoint:
                    logger.info("Resuming from checkpoint")
                    all_places = checkpoint.get('all_places', [])
                    processed_places = checkpoint.get('processed_places', [])
                    api_calls = checkpoint.get('api_calls', 0)
                    grid_points = checkpoint.get('grid_points', [])
                    current_grid_index = checkpoint.get('current_grid_index', 0)
                    center_location = checkpoint.get('center_location')
                    last_api_call_time = checkpoint.get('last_api_call_time', time.time())
                else:
                    logger.info("Starting new search")
                    all_places = []
                    processed_places = []
                    api_calls = 0
                    current_grid_index = 0
                    center_location = None
                    last_api_call_time = time.time()
                
                yield "Starting new search...\n"
                yield f"Using search radius: {radius} km\n"
                yield f"Using search density: {density} ({get_grid_size(density)}x{get_grid_size(density)} grid)\n"
                yield "Using 5km radius for each individual search point\n"
                yield "Fetching up to 60 results per search point\n"
                
                # Initialize Google Maps client
                gmaps = googlemaps.Client(key=api_key)
                
                if not center_location:
                    yield f"Converting location '{location_name}' to coordinates...\n"
                    try:
                        center_location = get_location_coordinates(gmaps, location_name)
                        logger.info(f"Found center coordinates: {center_location}")
                        yield f"Found center coordinates: {center_location}\n"
                        
                        # Create search grid
                        grid_points = create_search_grid(center_location[0], center_location[1], radius, density)
                        logger.info(f"Created search grid with {len(grid_points)} points")
                        yield f"Created search grid with {len(grid_points)} points\n"
                        
                    except Exception as e:
                        logger.error(f"Error getting coordinates: {str(e)}", exc_info=True)
                        yield f"Error: {str(e)}\n"
                        return
                
                yield f"Searching for {industry} businesses in the area...\n"
                
                # Search for places in each grid point
                for i in range(current_grid_index, len(grid_points)):
                    current_time = time.time()
                    time_since_last_call = current_time - last_api_call_time
                    
                    # Ensure we don't exceed 59 API calls per minute with more conservative delays
                    if time_since_last_call < 2.0:  # Less than 2 seconds since last call
                        sleep_time = 2.0 - time_since_last_call
                        logger.info(f"Rate limiting: waiting {sleep_time:.2f} seconds")
                        time.sleep(sleep_time)
                    
                    lat, lng = grid_points[i]
                    logger.info(f"Searching grid point {i+1}/{len(grid_points)} at ({lat}, {lng})")
                    yield f"Searching grid point {i+1}/{len(grid_points)} at ({lat}, {lng})...\n"
                    
                    try:
                        places = search_places(gmaps, (lat, lng), industry)
                        all_places.extend(places)
                        api_calls += 1
                        last_api_call_time = time.time()
                        logger.info(f"Found {len(places)} places at grid point {i+1}")
                        yield f"Found {len(places)} places at this point\n"
                        
                        # Update task progress
                        progress = (i + 1) / len(grid_points) * 50  # 50% for grid search
                        active_tasks[task_id]['progress'] = progress
                        active_tasks[task_id]['message'] = f"Searching grid point {i+1}/{len(grid_points)}"
                        
                        # Save checkpoint after each grid point
                        save_checkpoint({
                            'all_places': all_places,
                            'processed_places': processed_places,
                            'api_calls': api_calls,
                            'grid_points': grid_points,
                            'current_grid_index': i + 1,
                            'center_location': center_location,
                            'last_api_call_time': last_api_call_time
                        }, checkpoint_file)
                        
                        # Force garbage collection after each grid point
                        gc.collect()
                        
                        # Add additional delay between grid points
                        time.sleep(5)  # Added 5-second delay between grid points
                        
                    except Exception as e:
                        logger.error(f"Error searching grid point {i+1}: {str(e)}", exc_info=True)
                        yield f"Error searching grid point: {str(e)}\n"
                        # Save checkpoint on error
                        save_checkpoint({
                            'all_places': all_places,
                            'processed_places': processed_places,
                            'api_calls': api_calls,
                            'grid_points': grid_points,
                            'current_grid_index': i,
                            'center_location': center_location,
                            'last_api_call_time': last_api_call_time
                        }, checkpoint_file)
                        continue
                
                # Remove duplicates based on place_id
                unique_places = {place['place_id']: place for place in all_places}.values()
                logger.info(f"Total unique places found: {len(unique_places)}")
                yield f"Total unique places found: {len(unique_places)}\n"
                
                # Process each place
                businesses = []
                for i, place in enumerate(unique_places, 1):
                    current_time = time.time()
                    time_since_last_call = current_time - last_api_call_time
                    
                    # Ensure we don't exceed 59 API calls per minute with more conservative delays
                    if time_since_last_call < 2.0:  # Less than 2 seconds since last call
                        sleep_time = 2.0 - time_since_last_call
                        logger.info(f"Rate limiting: waiting {sleep_time:.2f} seconds")
                        time.sleep(sleep_time)
                    
                    place_id = place['place_id']
                    if place_id in processed_places:
                        continue
                        
                    place_name = place.get('name', 'Unknown')
                    logger.info(f"Processing {i}/{len(unique_places)}: {place_name}")
                    yield f"Processing {i}/{len(unique_places)}: {place_name}...\n"
                    
                    try:
                        details = get_place_details(gmaps, place_id)
                        if details:
                            businesses.append(details)
                        processed_places.append(place_id)
                        api_calls += 1
                        last_api_call_time = time.time()
                        
                        # Update task progress
                        progress = 50 + (i / len(unique_places) * 50)  # 50-100% for place details
                        active_tasks[task_id]['progress'] = progress
                        active_tasks[task_id]['message'] = f"Processing place {i}/{len(unique_places)}"
                        
                        # Save checkpoint after each place
                        save_checkpoint({
                            'all_places': all_places,
                            'processed_places': processed_places,
                            'api_calls': api_calls,
                            'grid_points': grid_points,
                            'current_grid_index': len(grid_points),
                            'center_location': center_location,
                            'last_api_call_time': last_api_call_time
                        }, checkpoint_file)
                        
                        # Force garbage collection after each business
                        gc.collect()
                        
                        # Add additional delay between place details
                        time.sleep(5)  # Added 5-second delay between place details
                        
                    except Exception as e:
                        logger.error(f"Error getting details for {place_name}: {str(e)}", exc_info=True)
                        yield f"Error getting details for {place_name}: {str(e)}\n"
                        # Save checkpoint on error
                        save_checkpoint({
                            'all_places': all_places,
                            'processed_places': processed_places,
                            'api_calls': api_calls,
                            'grid_points': grid_points,
                            'current_grid_index': len(grid_points),
                            'center_location': center_location,
                            'last_api_call_time': last_api_call_time
                        }, checkpoint_file)
                        continue
                
                # Save to Excel
                filename = f"{industry.replace(' ', '_')}_businesses.xlsx"
                file_path = save_to_excel(businesses, filename)
                
                # Clean up checkpoint file after successful completion
                cleanup_checkpoint(checkpoint_file)
                
                # Update task status
                active_tasks[task_id]['status'] = 'completed'
                active_tasks[task_id]['progress'] = 100
                active_tasks[task_id]['message'] = 'Search completed'
                task_results[task_id] = {
                    'filename': filename,
                    'businesses': len(businesses)
                }
                
                logger.info(f"Successfully completed search and saved results to {filename}")
                yield f"Success! Results saved to {filename}\n"
                yield f"DOWNLOAD_URL:/download/{filename}\n"
                
            except Exception as e:
                logger.error(f"Error in generate_updates: {str(e)}", exc_info=True)
                active_tasks[task_id]['status'] = 'error'
                active_tasks[task_id]['message'] = str(e)
                yield f"Error: {str(e)}\n"
                return
            
        return Response(stream_with_context(generate_updates()), mimetype='text/plain')
        
    except Exception as e:
        logger.error(f"Error in search: {str(e)}", exc_info=True)
        return Response(f"Error: {str(e)}\n", mimetype='text/plain', status=500)

@app.route('/task/<task_id>')
def get_task_status(task_id):
    """Get the status of a search task."""
    if task_id in active_tasks:
        return jsonify(active_tasks[task_id])
    elif task_id in task_results:
        return jsonify({
            'status': 'completed',
            'progress': 100,
            'message': 'Search completed',
            'result': task_results[task_id]
        })
    else:
        return jsonify({
            'status': 'not_found',
            'message': 'Task not found'
        }), 404

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(DOCUMENTS_DIR, filename)
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)


